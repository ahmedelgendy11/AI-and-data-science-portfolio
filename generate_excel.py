"""
Excel Mastery Workbook Generator
Creates an Excel workbook with multiple sheets for the Excel Mastery project.
Uses openpyxl to create a proper .xlsx file with formatting.
"""
import csv
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("openpyxl not installed. Installing...")
    import subprocess
    subprocess.check_call(["pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference

# ============================================================
# Load the sales data from CSV
# ============================================================
sales_data = []
with open("datasets/excel_sales_data.csv", "r", encoding="utf-8") as f:
    reader = csv.DictReader(f)
    for row in reader:
        sales_data.append(row)

print(f"Loaded {len(sales_data)} rows from excel_sales_data.csv")

# ============================================================
# Create the Excel Workbook
# ============================================================
wb = Workbook()

# Color palette
DARK_BLUE = "1B2A4A"
ACCENT_BLUE = "4A90D9"
ACCENT_GREEN = "27AE60"
ACCENT_ORANGE = "E67E22"
LIGHT_GRAY = "F5F6FA"
WHITE = "FFFFFF"
HEADER_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=12, bold=True, color=WHITE)
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color=DARK_BLUE)
SUBTITLE_FONT = Font(name="Calibri", size=12, bold=True, color=ACCENT_BLUE)
NORMAL_FONT = Font(name="Calibri", size=11)
ARABIC_FONT = Font(name="Calibri", size=11, color="333333")
BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD")
)

# ============================================================
# SHEET 1: Raw Sales Data
# ============================================================
ws_raw = wb.active
ws_raw.title = "Raw_Sales_Data"
ws_raw.sheet_properties.tabColor = DARK_BLUE

headers = ["sale_id", "date", "region", "sales_rep", "category", "product", "unit_price", "quantity", "discount_pct", "revenue", "cost", "profit", "marketing_spend"]
for col, header in enumerate(headers, 1):
    cell = ws_raw.cell(row=1, column=col, value=header.upper().replace("_", " "))
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER

for row_idx, row_data in enumerate(sales_data, 2):
    for col_idx, header in enumerate(headers, 1):
        value = row_data[header]
        # Convert numeric columns
        if header in ["unit_price", "revenue", "cost", "profit", "marketing_spend", "discount_pct"]:
            value = float(value)
        elif header == "quantity":
            value = int(value)
        cell = ws_raw.cell(row=row_idx, column=col_idx, value=value)
        cell.font = NORMAL_FONT
        cell.border = BORDER
        if header in ["unit_price", "revenue", "cost", "profit", "marketing_spend"]:
            cell.number_format = '#,##0.00'
        elif header == "discount_pct":
            cell.number_format = '0%'

# Auto-adjust column widths
for col in range(1, len(headers) + 1):
    ws_raw.column_dimensions[get_column_letter(col)].width = 16

# Freeze header row
ws_raw.freeze_panes = "A2"

# ============================================================
# SHEET 2: Challenge Instructions (Arabic)
# ============================================================
ws_instructions = wb.create_sheet("Instructions")
ws_instructions.sheet_properties.tabColor = ACCENT_BLUE

instructions = [
    ("", ""),
    ("DATA ANALYST MASTERY - EXCEL CHALLENGES", ""),
    ("", ""),
    ("Challenge 1: Dynamic Array Formulas", ""),
    ("", "In the 'Dynamic_Arrays' sheet, use modern Excel formulas to:"),
    ("", "1. Use UNIQUE to extract all unique regions"),
    ("", "2. Use FILTER to show only 'Software' sales with revenue > $5000"),
    ("", "3. Use SORT to sort products by total revenue descending"),
    ("", "4. Use SORTBY to sort sales reps by their total profit"),
    ("", "5. Use LET to create a formula that calculates profit margin % per region"),
    ("", "6. Use XLOOKUP to find the top sales rep for each region"),
    ("", ""),
    ("Challenge 2: SUMIFS, COUNTIFS & Conditional Analysis", ""),
    ("", "In the 'Conditional_Analysis' sheet:"),
    ("", "1. Calculate total revenue per region using SUMIFS"),
    ("", "2. Count number of transactions per category per region using COUNTIFS"),
    ("", "3. Calculate average discount per sales rep using AVERAGEIFS"),
    ("", "4. Find max revenue transaction per region"),
    ("", "5. Calculate YoY growth rate for each quarter"),
    ("", ""),
    ("Challenge 3: Pivot Table Analysis", ""),
    ("", "Create Pivot Tables in the 'Pivot_Analysis' sheet:"),
    ("", "1. Revenue by Region and Category (with grand totals)"),
    ("", "2. Monthly revenue trend with category breakdown"),
    ("", "3. Sales rep performance ranking by profit"),
    ("", "4. Add Slicers for Region, Category, and Date"),
    ("", ""),
    ("Challenge 4: Statistical Analysis", ""),
    ("", "In the 'Statistics' sheet:"),
    ("", "1. Calculate descriptive statistics (mean, median, std, skewness)"),
    ("", "2. Run a linear regression: Revenue vs Marketing Spend"),
    ("", "3. Calculate R-squared and interpret the result"),
    ("", "4. Create a scatter plot with trendline"),
    ("", "5. Perform a correlation analysis between all numeric columns"),
    ("", ""),
    ("Challenge 5: Power Query (Data Transformation)", ""),
    ("", "Use Power Query (Get & Transform Data) to:"),
    ("", "1. Load the Raw_Sales_Data into Power Query Editor"),
    ("", "2. Add a 'Quarter' column derived from the date"),
    ("", "3. Add a 'Profit Margin %' calculated column"),
    ("", "4. Group by Region and Category to get sum of revenue"),
    ("", "5. Unpivot the revenue/cost/profit columns for analysis"),
    ("", "6. Create a calendar/date dimension table"),
    ("", ""),
    ("Challenge 6: Power Pivot & DAX Measures", ""),
    ("", "Enable Power Pivot and create a Data Model:"),
    ("", "1. Create relationships between Sales and a Date table"),
    ("", "2. Write DAX: Total Revenue = SUM(Sales[revenue])"),
    ("", "3. Write DAX: YTD Revenue = TOTALYTD([Total Revenue], Dates[Date])"),
    ("", "4. Write DAX: YoY Growth = DIVIDE([Total Revenue] - [PY Revenue], [PY Revenue])"),
    ("", "5. Write DAX: Running Total = CALCULATE([Total Revenue], FILTER(...))"),
    ("", "6. Create a PivotTable using the Data Model measures"),
    ("", ""),
    ("Challenge 7: VBA Macro Automation", ""),
    ("", "Write VBA code (Alt+F11) to:"),
    ("", "1. Create a macro that formats any data table with headers, borders, and alternating row colors"),
    ("", "2. Create a macro that generates a summary report on a new sheet"),
    ("", "3. Create a macro that exports the Dashboard sheet as a PDF"),
    ("", "4. Add buttons to trigger these macros"),
    ("", ""),
    ("Challenge 8: Interactive Dashboard Design", ""),
    ("", "In the 'Dashboard' sheet, create a professional dashboard with:"),
    ("", "1. KPI Cards: Total Revenue, Total Profit, Avg Profit Margin, Total Orders"),
    ("", "2. Bar Chart: Revenue by Region"),
    ("", "3. Line Chart: Monthly Revenue Trend"),
    ("", "4. Pie Chart: Revenue by Category"),
    ("", "5. Table: Top 10 Sales Reps by Revenue"),
    ("", "6. Slicers for Region, Category, Year"),
    ("", "7. Consistent color theme and professional formatting"),
]

for row_idx, (col_a, col_b) in enumerate(instructions, 1):
    cell_a = ws_instructions.cell(row=row_idx, column=1, value=col_a)
    cell_b = ws_instructions.cell(row=row_idx, column=2, value=col_b)
    
    if col_a and not col_b:  # Title rows
        if "MASTERY" in col_a:
            cell_a.font = TITLE_FONT
        else:
            cell_a.font = SUBTITLE_FONT
    else:
        cell_a.font = ARABIC_FONT
        cell_b.font = ARABIC_FONT

ws_instructions.column_dimensions["A"].width = 55
ws_instructions.column_dimensions["B"].width = 80

# ============================================================
# SHEET 3: Dynamic Arrays (workspace for Challenge 1)
# ============================================================
ws_dynamic = wb.create_sheet("Dynamic_Arrays")
ws_dynamic.sheet_properties.tabColor = ACCENT_GREEN

# Add section headers
sections = [
    (1, "DYNAMIC ARRAY FORMULAS WORKSPACE"),
    (3, "1. UNIQUE - Extract unique regions (write formula in B4):"),
    (10, "2. FILTER - Software sales > $5000 (write formula in B11):"),
    (22, "3. SORT - Products sorted by revenue (write formula in B23):"),
    (34, "4. SORTBY - Sales reps sorted by profit (write formula in B35):"),
    (46, "5. LET - Profit margin per region (write formula in B47):"),
    (56, "6. XLOOKUP - Top rep per region (write formula in B57):"),
]

for row, text in sections:
    cell = ws_dynamic.cell(row=row, column=1, value=text)
    if row == 1:
        cell.font = TITLE_FONT
    else:
        cell.font = SUBTITLE_FONT

ws_dynamic.column_dimensions["A"].width = 60
ws_dynamic.column_dimensions["B"].width = 25

# ============================================================
# SHEET 4: Conditional Analysis (workspace for Challenge 2)
# ============================================================
ws_cond = wb.create_sheet("Conditional_Analysis")
ws_cond.sheet_properties.tabColor = ACCENT_ORANGE

# Region headers for SUMIFS
ws_cond.cell(row=1, column=1, value="CONDITIONAL ANALYSIS WORKSPACE").font = TITLE_FONT
ws_cond.cell(row=3, column=1, value="1. Total Revenue by Region (SUMIFS)").font = SUBTITLE_FONT

regions = ["North", "South", "East", "West", "Central"]
ws_cond.cell(row=4, column=1, value="Region").font = HEADER_FONT
ws_cond.cell(row=4, column=1).fill = HEADER_FILL
ws_cond.cell(row=4, column=2, value="Total Revenue").font = HEADER_FONT
ws_cond.cell(row=4, column=2).fill = HEADER_FILL
for i, region in enumerate(regions, 5):
    ws_cond.cell(row=i, column=1, value=region).font = NORMAL_FONT
    ws_cond.cell(row=i, column=2, value="< Write SUMIFS formula here >").font = Font(color="999999", italic=True)

ws_cond.cell(row=12, column=1, value="2. Count Transactions per Category per Region (COUNTIFS)").font = SUBTITLE_FONT
categories = ["Software", "Hardware", "Services", "Accessories"]
ws_cond.cell(row=13, column=1, value="Category").font = HEADER_FONT
ws_cond.cell(row=13, column=1).fill = HEADER_FILL
for j, region in enumerate(regions, 2):
    ws_cond.cell(row=13, column=j, value=region).font = HEADER_FONT
    ws_cond.cell(row=13, column=j).fill = HEADER_FILL
for i, cat in enumerate(categories, 14):
    ws_cond.cell(row=i, column=1, value=cat).font = NORMAL_FONT

ws_cond.cell(row=20, column=1, value="3. Average Discount per Sales Rep (AVERAGEIFS)").font = SUBTITLE_FONT

ws_cond.column_dimensions["A"].width = 55
ws_cond.column_dimensions["B"].width = 25

# ============================================================
# SHEET 5: Pivot Analysis (placeholder for Challenge 3)
# ============================================================
ws_pivot = wb.create_sheet("Pivot_Analysis")
ws_pivot.sheet_properties.tabColor = "9B59B6"
ws_pivot.cell(row=1, column=1, value="PIVOT TABLE ANALYSIS").font = TITLE_FONT
ws_pivot.cell(row=3, column=1, value="Instructions: Create Pivot Tables from Raw_Sales_Data here.").font = ARABIC_FONT
ws_pivot.cell(row=4, column=1, value="1. Insert > PivotTable > Select Raw_Sales_Data range").font = ARABIC_FONT
ws_pivot.cell(row=5, column=1, value="2. Place Pivot Tables starting from row 7").font = ARABIC_FONT
ws_pivot.column_dimensions["A"].width = 60

# ============================================================
# SHEET 6: Statistics (workspace for Challenge 4)
# ============================================================
ws_stats = wb.create_sheet("Statistics")
ws_stats.sheet_properties.tabColor = "E74C3C"
ws_stats.cell(row=1, column=1, value="STATISTICAL ANALYSIS WORKSPACE").font = TITLE_FONT
ws_stats.cell(row=3, column=1, value="1. Descriptive Statistics").font = SUBTITLE_FONT

stat_labels = ["Mean", "Median", "Mode", "Std Dev", "Variance", "Skewness", "Kurtosis", "Min", "Max", "Count"]
stat_columns = ["Revenue", "Profit", "Unit Price", "Marketing Spend"]
ws_stats.cell(row=4, column=1, value="Statistic").font = HEADER_FONT
ws_stats.cell(row=4, column=1).fill = HEADER_FILL
for j, col_name in enumerate(stat_columns, 2):
    ws_stats.cell(row=4, column=j, value=col_name).font = HEADER_FONT
    ws_stats.cell(row=4, column=j).fill = HEADER_FILL
for i, label in enumerate(stat_labels, 5):
    ws_stats.cell(row=i, column=1, value=label).font = NORMAL_FONT

ws_stats.cell(row=17, column=1, value="2. Linear Regression: Revenue vs Marketing Spend").font = SUBTITLE_FONT
ws_stats.cell(row=18, column=1, value="Use Data Analysis ToolPak > Regression").font = ARABIC_FONT
ws_stats.cell(row=20, column=1, value="3. Correlation Matrix").font = SUBTITLE_FONT

ws_stats.column_dimensions["A"].width = 30
for col in range(2, 6):
    ws_stats.column_dimensions[get_column_letter(col)].width = 18

# ============================================================
# SHEET 7: Dashboard (workspace for Challenge 8)
# ============================================================
ws_dash = wb.create_sheet("Dashboard")
ws_dash.sheet_properties.tabColor = "2ECC71"
ws_dash.sheet_view.showGridLines = False

# KPI section headers
kpi_labels = ["TOTAL REVENUE", "TOTAL PROFIT", "AVG PROFIT MARGIN", "TOTAL ORDERS"]
kpi_colors = ["1B6B93", "27AE60", "E67E22", "8E44AD"]

for i, (label, color) in enumerate(zip(kpi_labels, kpi_colors)):
    col_start = i * 4 + 1
    cell = ws_dash.cell(row=2, column=col_start, value=label)
    cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center")
    
    value_cell = ws_dash.cell(row=3, column=col_start, value="< Formula >")
    value_cell.font = Font(name="Calibri", size=18, bold=True, color=color)
    value_cell.alignment = Alignment(horizontal="center")

ws_dash.cell(row=1, column=1, value="SALES PERFORMANCE DASHBOARD").font = Font(name="Calibri", size=20, bold=True, color=DARK_BLUE)

ws_dash.cell(row=6, column=1, value="[ Place Bar Chart: Revenue by Region here ]").font = Font(color="999999", size=12, italic=True)
ws_dash.cell(row=6, column=8, value="[ Place Pie Chart: Revenue by Category here ]").font = Font(color="999999", size=12, italic=True)
ws_dash.cell(row=20, column=1, value="[ Place Line Chart: Monthly Revenue Trend here ]").font = Font(color="999999", size=12, italic=True)
ws_dash.cell(row=20, column=8, value="[ Top 10 Sales Reps Table here ]").font = Font(color="999999", size=12, italic=True)

for col in range(1, 17):
    ws_dash.column_dimensions[get_column_letter(col)].width = 14

# ============================================================
# SHEET 8: VBA Instructions
# ============================================================
ws_vba = wb.create_sheet("VBA_Guide")
ws_vba.sheet_properties.tabColor = "34495E"
ws_vba.cell(row=1, column=1, value="VBA MACRO GUIDE").font = TITLE_FONT

vba_content = [
    (3, "How to open VBA Editor: Press Alt + F11"),
    (4, ""),
    (5, "MACRO 1: Auto-Format Table"),
    (6, "Sub FormatTable()"),
    (7, "    ' Select the data range"),
    (8, "    ' Apply header formatting (bold, background color)"),
    (9, "    ' Apply borders to all cells"),
    (10, "    ' Apply alternating row colors"),
    (11, "    ' Auto-fit column widths"),
    (12, "End Sub"),
    (13, ""),
    (14, "MACRO 2: Generate Summary Report"),
    (15, "Sub GenerateReport()"),
    (16, "    ' Create a new sheet named 'Summary_Report'"),
    (17, "    ' Calculate total revenue, profit, orders per region"),
    (18, "    ' Add charts automatically"),
    (19, "    ' Format the report professionally"),
    (20, "End Sub"),
    (21, ""),
    (22, "MACRO 3: Export to PDF"),
    (23, "Sub ExportDashboardPDF()"),
    (24, "    ' Select the Dashboard sheet"),
    (25, "    ' Set print area"),
    (26, "    ' Export as PDF to the same folder"),
    (27, "    ' Show confirmation message"),
    (28, "End Sub"),
    (29, ""),
    (30, "TIP: Record a macro first (View > Macros > Record), then edit the generated code to learn VBA syntax!"),
]

for row, text in vba_content:
    cell = ws_vba.cell(row=row, column=1, value=text)
    if text.startswith("MACRO") or text.startswith("TIP"):
        cell.font = SUBTITLE_FONT
    elif text.startswith("Sub") or text.startswith("End Sub"):
        cell.font = Font(name="Consolas", size=11, bold=True, color="2E86C1")
    elif text.startswith("    '"):
        cell.font = Font(name="Consolas", size=11, color=ACCENT_GREEN)
    else:
        cell.font = NORMAL_FONT

ws_vba.column_dimensions["A"].width = 80

# ============================================================
# Save the workbook
# ============================================================
output_path = "project_4_excel_mastery.xlsx"
wb.save(output_path)
print(f"\nExcel workbook saved: {output_path}")
print(f"Sheets: {wb.sheetnames}")
print("Done!")
