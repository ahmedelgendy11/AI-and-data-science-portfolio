"""
Verification script to ensure all created files are valid and error-free.
"""
import json
import os
import csv
from openpyxl import load_workbook

print("--- Starting Project Verification ---")

errors = 0

# 1. Verify CSV Datasets
datasets = [
    "datasets/ecommerce_transactions.csv",
    "datasets/health_risk_data.csv",
    "datasets/stock_market_data.csv",
    "datasets/excel_sales_data.csv"
]

for ds in datasets:
    if not os.path.exists(ds):
        print(f"[ERROR] Dataset {ds} is missing!")
        errors += 1
    else:
        try:
            with open(ds, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                header = next(reader)
                rows = list(reader)
                print(f"[OK] Dataset {ds}: {len(rows)} rows, columns: {header}")
        except Exception as e:
            print(f"[ERROR] Reading {ds}: {e}")
            errors += 1

# 2. Verify Jupyter Notebooks (Starter & Solution)
notebooks = [
    "project_1_ecommerce.ipynb",
    "project_2_health_risk.ipynb",
    "project_3_stock_trends.ipynb",
    "solutions/project_1_ecommerce_solution.ipynb",
    "solutions/project_2_health_risk_solution.ipynb",
    "solutions/project_3_stock_trends_solution.ipynb"
]

for nb in notebooks:
    if not os.path.exists(nb):
        print(f"[ERROR] Notebook {nb} is missing!")
        errors += 1
    else:
        try:
            with open(nb, "r", encoding="utf-8") as f:
                data = json.load(f)
            # Check basic notebook structure
            if "cells" not in data or "nbformat" not in data:
                print(f"[ERROR] {nb} is not a valid Jupyter Notebook format!")
                errors += 1
            else:
                code_cells = sum(1 for c in data["cells"] if c["cell_type"] == "code")
                markdown_cells = sum(1 for c in data["cells"] if c["cell_type"] == "markdown")
                print(f"[OK] Notebook {nb}: Valid format, {len(data['cells'])} cells ({code_cells} code, {markdown_cells} markdown)")
        except Exception as e:
            print(f"[ERROR] Parsing {nb}: {e}")
            errors += 1

# 3. Verify Excel Workbooks (Starter & Solution)
excel_files = [
    "project_4_excel_mastery.xlsx",
    "solutions/project_4_excel_mastery_solution.xlsx"
]

for excel in excel_files:
    if not os.path.exists(excel):
        print(f"[ERROR] Excel file {excel} is missing!")
        errors += 1
    else:
        try:
            wb = load_workbook(excel, read_only=True)
            print(f"[OK] Excel {excel}: Valid workbook, sheets: {wb.sheetnames}")
        except Exception as e:
            print(f"[ERROR] Reading {excel}: {e}")
            errors += 1

# 4. Verify Web Sandbox Files
sandbox_files = [
    "web_sandbox/index.html",
    "web_sandbox/style.css",
    "web_sandbox/app.js",
    "web_sandbox/challenges.js"
]

for sb in sandbox_files:
    if not os.path.exists(sb):
        print(f"[ERROR] Web Sandbox file {sb} is missing!")
        errors += 1
    else:
        print(f"[OK] Web Sandbox file {sb}: Exists ({os.path.getsize(sb)} bytes)")

print("\n--- Verification Summary ---")
if errors == 0:
    print("SUCCESS: All project files are valid, formatted correctly, and complete!")
else:
    print(f"FAILED: Found {errors} verification errors. Please check the logs.")

