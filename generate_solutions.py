"""
Python script to generate solution notebooks for Phase 5 of the Data Analyst Mastery Suite.
Replaces all placeholders with the correct code and executes the notebooks to pre-populate outputs.
"""
import json
import os
import subprocess
from openpyxl.utils import get_column_letter


os.makedirs("solutions", exist_ok=True)

# ============================================================
# SOLVE PROJECT 1: E-COMMERCE
# ============================================================
print("Solving Project 1 (E-Commerce)...")
with open("project_1_ecommerce.ipynb", "r", encoding="utf-8") as f:
    nb1 = json.load(f)

for cell in nb1["cells"]:
    if cell["cell_type"] == "code":
        source = cell["source"]
        new_source = []
        for line in source:
            # Exercise 1
            if "unique_customers = ___" in line:
                line = line.replace("unique_customers = ___", "unique_customers = df['customer_id'].nunique()")
            if "unique_products = ___" in line:
                line = line.replace("unique_products = ___", "unique_products = df['product_name'].nunique()")
            if "num_categories = ___" in line:
                line = line.replace("num_categories = ___", "num_categories = df['category'].nunique()")
            if "category_names = ___" in line:
                line = line.replace("category_names = ___", "category_names = df['category'].unique()")
            if "unique_countries = ___" in line:
                line = line.replace("unique_countries = ___", "unique_countries = df['country'].nunique()")
            
            # Exercise 2
            if "missing_counts = ___" in line:
                line = line.replace("missing_counts = ___", "missing_counts = df.isnull().sum()")
            if "missing_pct = ___" in line:
                line = line.replace("missing_pct = ___", "missing_pct = (df.isnull().sum() / len(df)) * 100")
            
            # Exercise 3
            if "___,  # TODO: Add format string for YYYY-MM-DD" in line:
                line = line.replace("___,  # TODO: Add format string for YYYY-MM-DD (e.g., '%Y-%m-%d')", "'%Y-%m-%d',")
            if "___,  # TODO: Add format string for DD/MM/YYYY" in line:
                line = line.replace("___,  # TODO: Add format string for DD/MM/YYYY (e.g., '%d/%m/%Y')", "'%d/%m/%Y',")
            if "___,  # TODO: Add format string for MM-DD-YYYY" in line:
                line = line.replace("___,  # TODO: Add format string for MM-DD-YYYY (e.g., '%m-%d-%Y')", "'%m-%d-%Y',")
            if "df['order_date'] = ___" in line:
                line = line.replace("df['order_date'] = ___", "df['order_date'] = df['order_date'].apply(parse_mixed_date)")
            
            # Exercise 4
            if "num_duplicates = ___" in line:
                line = line.replace("num_duplicates = ___", "num_duplicates = df.duplicated().sum()")
            if "duplicated_rows = ___" in line:
                line = line.replace("duplicated_rows = ___", "duplicated_rows = df[df.duplicated(keep=False)].sort_values('order_id').head(10)")
            if "df = ___" in line:
                line = line.replace("df = ___", "df = df.drop_duplicates(keep='first').reset_index(drop=True)")
            
            # Exercise 5
            if "df['country'] = ___" in line:
                line = line.replace("df['country'] = ___", "df['country'] = df['country'].fillna('Unknown')")
            if "most_common_payment = ___" in line:
                line = line.replace("most_common_payment = ___", "most_common_payment = df['payment_method'].mode()[0]")
            if "df['payment_method'] = ___" in line:
                line = line.replace("df['payment_method'] = ___", "df['payment_method'] = df['payment_method'].fillna(most_common_payment)")
            
            # Exercise 6
            if "negative_qty = ___" in line:
                line = line.replace("negative_qty = ___", "negative_qty = df[df['quantity'] < 0]")
            if "df_returns = negative_qty.copy()" in line:
                pass
            if "df = ___  # use df[df['quantity'] > 0]" in line:
                line = line.replace("df = ___  # use df[df['quantity'] > 0].reset_index(drop=True)", "df = df[df['quantity'] > 0].reset_index(drop=True)")
            
            # Exercise 7
            if "revenue_by_category = ___" in line:
                line = line.replace("revenue_by_category = ___", "revenue_by_category = df.groupby('category')['total_amount'].sum().sort_values(ascending=False)")
            if "___  # use revenue_by_category.plot" in line:
                line = line.replace("___  # use revenue_by_category.plot(kind='bar', ax=ax, color=sns.color_palette('viridis', len(revenue_by_category)))", "revenue_by_category.plot(kind='bar', ax=ax, color=sns.color_palette('viridis', len(revenue_by_category)))")
            
            # Exercise 8
            if "df['order_month'] = ___" in line:
                line = line.replace("df['order_month'] = ___", "df['order_month'] = df['order_date'].dt.to_period('M')")
            if "monthly_revenue = ___" in line:
                line = line.replace("monthly_revenue = ___", "monthly_revenue = df.groupby('order_month')['total_amount'].sum()")
            if "___  # use monthly_revenue.plot" in line:
                line = line.replace("___  # use monthly_revenue.plot(kind='line', ax=ax, marker='o', linewidth=2, color='#2196F3')", "monthly_revenue.plot(kind='line', ax=ax, marker='o', linewidth=2, color='#2196F3')")
            
            # Exercise 9
            if "revenue_by_product = ___" in line:
                line = line.replace("revenue_by_product = ___", "revenue_by_product = df.groupby('product_name')['total_amount'].sum()")
            if "top_10_products = ___" in line:
                line = line.replace("top_10_products = ___", "top_10_products = revenue_by_product.nlargest(10)")
            if "___  # use top_10_products.sort_values" in line:
                line = line.replace("___  # use top_10_products.sort_values().plot(kind='barh', ax=ax, color=sns.color_palette('coolwarm', 10))", "top_10_products.sort_values().plot(kind='barh', ax=ax, color=sns.color_palette('coolwarm', 10))")
            
            # Exercise 10
            if "revenue_by_country = ___" in line:
                line = line.replace("revenue_by_country = ___", "revenue_by_country = df.groupby('country')['total_amount'].sum().sort_values(ascending=False)")
            if "___  # use plot_data.plot" in line:
                line = line.replace("___  # use plot_data.plot(kind='pie', ax=ax, autopct='%1.1f%%', colors=colors, startangle=90, textprops={'fontsize': 11})", "plot_data.plot(kind='pie', ax=ax, autopct='%1.1f%%', colors=colors, startangle=90, textprops={'fontsize': 11})")
            
            # Exercise 11
            if "payment_counts = ___" in line:
                line = line.replace("payment_counts = ___", "payment_counts = df['payment_method'].value_counts()")
            if "payment_revenue = ___" in line:
                line = line.replace("payment_revenue = ___", "payment_revenue = df.groupby('payment_method')['total_amount'].sum().sort_values(ascending=False)")
            if "___  # use payment_counts.plot" in line:
                line = line.replace("___  # use payment_counts.plot(kind='bar', ax=ax1, color=sns.color_palette('pastel'))", "payment_counts.plot(kind='bar', ax=ax1, color=sns.color_palette('pastel'))")
            if "___  # use payment_revenue.plot" in line:
                line = line.replace("___  # use payment_revenue.plot(kind='bar', ax=ax2, color=sns.color_palette('muted'))", "payment_revenue.plot(kind='bar', ax=ax2, color=sns.color_palette('muted'))")
            
            # Exercise 12
            if "last_purchase = ___" in line:
                line = line.replace("last_purchase = ___", "last_purchase = df.groupby('customer_id')['order_date'].max()")
            if "recency = ___" in line:
                line = line.replace("recency = ___", "recency = (reference_date - last_purchase).dt.days")
            
            # Exercise 13
            if "frequency = ___" in line:
                line = line.replace("frequency = ___", "frequency = df.groupby('customer_id')['order_id'].nunique()")
            
            # Exercise 14
            if "monetary = ___" in line:
                line = line.replace("monetary = ___", "monetary = df.groupby('customer_id')['total_amount'].sum()")
            
            # Exercise 15
            if "rfm['r_score'] = ___" in line:
                line = line.replace("rfm['r_score'] = ___", "rfm['r_score'] = pd.qcut(rfm['recency'], 5, labels=[5, 4, 3, 2, 1], duplicates='drop')")
            if "rfm['f_score'] = ___" in line:
                line = line.replace("rfm['f_score'] = ___", "rfm['f_score'] = pd.qcut(rfm['frequency'].rank(method='first'), 5, labels=[1, 2, 3, 4, 5], duplicates='drop')")
            if "rfm['m_score'] = ___" in line:
                line = line.replace("rfm['m_score'] = ___", "rfm['m_score'] = pd.qcut(rfm['monetary'], 5, labels=[1, 2, 3, 4, 5], duplicates='drop')")
            if "rfm['rfm_score'] = ___" in line:
                line = line.replace("rfm['rfm_score'] = ___", "rfm['rfm_score'] = rfm['r_score'].astype(str) + rfm['f_score'].astype(str) + rfm['m_score'].astype(str)")
            
            # Exercise 16
            if "elif ___:" in line and "Loyal" in line:
                line = line.replace("elif ___:                        # TODO: Condition for Loyal Customers (f >= 4)", "elif f >= 4:")
            if "elif ___:" in line and "Potential" in line:
                line = line.replace("elif ___:                        # TODO: Condition for Potential Loyalists (r >= 4 and f >= 2)", "elif r >= 4 and f >= 2:")
            if "elif ___:" in line and "At Risk" in line:
                line = line.replace("elif ___:                        # TODO: Condition for At Risk (r <= 2 and f >= 3)", "elif r <= 2 and f >= 3:")
            if "elif ___:" in line and "Lost" in line:
                line = line.replace("elif ___:                        # TODO: Condition for Lost (r <= 2 and f <= 2)", "elif r <= 2 and f <= 2:")
            if "rfm['segment'] = ___" in line:
                line = line.replace("rfm['segment'] = ___", "rfm['segment'] = rfm.apply(segment_customer, axis=1)")
            
            # Exercise 17
            if "segment_counts = ___" in line:
                line = line.replace("segment_counts = ___", "segment_counts = rfm['segment'].value_counts()")
            if "avg_monetary = ___" in line:
                line = line.replace("avg_monetary = ___", "avg_monetary = rfm.groupby('segment')['monetary'].mean().sort_values(ascending=False)")
            
            # Exercise 18
            if "df['first_purchase'] = ___" in line:
                line = line.replace("df['first_purchase'] = ___", "df['first_purchase'] = df.groupby('customer_id')['order_date'].transform('min')")
            if "df['cohort_month'] = ___" in line:
                line = line.replace("df['cohort_month'] = ___", "df['cohort_month'] = df['first_purchase'].dt.to_period('M')")
            
            # Exercise 19
            if "df['cohort_index'] = ___" in line:
                line = line.replace("df['cohort_index'] = ___", "df['cohort_index'] = (order_year - cohort_year) * 12 + (order_month_num - cohort_month_num)")
            
            # Exercise 20
            if "cohort_data = ___" in line:
                line = line.replace("cohort_data = ___", "cohort_data = df.groupby(['cohort_month', 'cohort_index'])['customer_id'].nunique().reset_index()")
            if "cohort_table = ___" in line:
                line = line.replace("cohort_table = ___", "cohort_table = cohort_data.pivot(index='cohort_month', columns='cohort_index', values='customer_count')")
            if "retention_table = ___" in line:
                line = line.replace("retention_table = ___", "retention_table = cohort_table.divide(cohort_sizes, axis=0) * 100")
            
            # Exercise 21 (Heatmap commented code)
            if "___  # use sns.heatmap(" in line:
                line = ""
            if "#     retention_table," in line:
                line = line.replace("#     retention_table,", "    retention_table,")
            if "#     annot=True," in line:
                line = line.replace("#     annot=True,", "    annot=True,")
            if "#     fmt='.0f'," in line:
                line = line.replace("#     fmt='.0f',", "    fmt='.0f',")
            if "#     cmap='YlOrRd_r'," in line:
                line = line.replace("#     cmap='YlOrRd_r',", "    cmap='YlOrRd_r',")
            if "#     ax=ax," in line:
                line = line.replace("#     ax=ax,", "    ax=ax,")
            if "#     vmin=0," in line:
                line = line.replace("#     vmin=0,", "    vmin=0,")
            if "#     vmax=100," in line:
                line = line.replace("#     vmax=100,", "    vmax=100,")
            if "#     linewidths=0.5," in line:
                line = line.replace("#     linewidths=0.5,", "    linewidths=0.5,")
            if "#     cbar_kws={'label': 'معدل الاحتفاظ (%)'}" in line:
                line = line.replace("#     cbar_kws={'label': 'معدل الاحتفاظ (%)'}", "    cbar_kws={'label': 'معدل الاحتفاظ (%)'}")
            if "# )" in line and "cbar_kws" not in line:
                line = line.replace("# )", ")")
            
            new_source.append(line)
        cell["source"] = new_source

with open("solutions/project_1_ecommerce_solution.ipynb", "w", encoding="utf-8") as f:
    json.dump(nb1, f, indent=2, ensure_ascii=False)


# ============================================================
# SOLVE PROJECT 2: HEALTH RISK (Already solved, just copy)
# ============================================================
print("Solving Project 2 (Health Risk)...")
with open("project_2_health_risk.ipynb", "r", encoding="utf-8") as f:
    nb2 = json.load(f)
with open("solutions/project_2_health_risk_solution.ipynb", "w", encoding="utf-8") as f:
    json.dump(nb2, f, indent=2, ensure_ascii=False)


# ============================================================
# SOLVE PROJECT 3: STOCK MARKET
# ============================================================
print("Solving Project 3 (Stock Market)...")
with open("project_3_stock_trends.ipynb", "r", encoding="utf-8") as f:
    nb3 = json.load(f)

for cell in nb3["cells"]:
    if cell["cell_type"] == "code":
        source = cell["source"]
        new_source = []
        for line in source:
            # Exercise 1.1
            if "earliest_date = ___" in line:
                line = line.replace("earliest_date = ___", "earliest_date = df['date'].min()")
            if "latest_date = ___" in line:
                line = line.replace("latest_date = ___", "latest_date = df['date'].max()")
            if "tickers = ___" in line:
                line = line.replace("tickers = ___", "tickers = df['ticker'].unique()")
            if "trading_days = ___" in line:
                line = line.replace("trading_days = ___", "trading_days = df.groupby('ticker')['date'].nunique()")
            
            # Exercise 1.2
            if "summary_stats = ___" in line:
                line = line.replace("summary_stats = ___", "summary_stats = df.groupby('ticker')['close'].describe()")
            
            # Exercise 2.1
            if "stock_data = ___  # Filter data for this ticker" in line:
                line = line.replace("stock_data = ___  # Filter data for this ticker", "stock_data = df[df['ticker'] == ticker].copy()")
            if "plt.plot(___, ___, label=ticker)" in line:
                line = line.replace("plt.plot(___, ___, label=ticker)  # Plot date vs close", "plt.plot(stock_data['date'], stock_data['close'], label=ticker)")
            
            # Exercise 2.2
            if "stock_data = ___  # Filter data for this ticker" in line and "2.2" in "".join(source):
                pass # Handled below
            if "axes[i].plot(___, ___, label=col" in line:
                line = line.replace("axes[i].plot(___, ___, label=col, color=color, alpha=0.7)  # Plot date vs price column", "axes[i].plot(stock_data['date'], stock_data[col], label=col, color=color, alpha=0.7)")
            
            # Exercise 2.3
            if "axes[i].fill_between(___, ___, alpha=0.4" in line:
                line = line.replace("axes[i].fill_between(___, ___, alpha=0.4, color=palette[i])  # Fill area chart for volume", "axes[i].fill_between(stock_data['date'], stock_data['volume'], alpha=0.4, color=palette[i])")
            
            # General ticker filter
            if "stock_data = ___" in line:
                line = line.replace("stock_data = ___", "stock_data = df[df['ticker'] == ticker].copy()")
            
            # Exercise 3.1
            if "df = df.sort_values(by=[___, ___])" in line:
                line = line.replace("df = df.sort_values(by=[___, ___]).reset_index(drop=True)", "df = df.sort_values(by=['ticker', 'date']).reset_index(drop=True)")
            if "df['daily_return'] = df.groupby(___)[___].pct_change()" in line:
                line = line.replace("df['daily_return'] = df.groupby(___)[___].pct_change()", "df['daily_return'] = df.groupby('ticker')['close'].pct_change()")
            
            # Exercise 3.2
            if "stock_returns = ___  # Get daily_return" in line:
                line = line.replace("stock_returns = ___  # Get daily_return for this ticker, drop NaN", "stock_returns = df[df['ticker'] == ticker]['daily_return'].dropna()")
            if "axes[i].hist(___, bins=50" in line:
                line = line.replace("axes[i].hist(___, bins=50, edgecolor='black', alpha=0.7)  # Plot histogram", "axes[i].hist(stock_returns, bins=50, edgecolor='black', alpha=0.7)")
            
            # Exercise 3.3
            if "cumulative = ___  # (1 + daily_return)" in line:
                line = line.replace("cumulative = ___  # (1 + daily_return).cumprod() - 1", "cumulative = (1 + stock_data['daily_return']).cumprod() - 1")
            
            # Exercise 3.4
            if "annualized_vol = ___  # Multiply daily_vol" in line:
                line = line.replace("annualized_vol = ___  # Multiply daily_vol by sqrt(252)", "annualized_vol = daily_vol * np.sqrt(252)")
            
            # Exercise 4.1
            if "aapl['SMA_20'] = ___" in line:
                line = line.replace("aapl['SMA_20'] = ___  # rolling mean with window=20", "aapl['SMA_20'] = aapl['close'].rolling(window=20).mean()")
            if "aapl['SMA_50'] = ___" in line:
                line = line.replace("aapl['SMA_50'] = ___  # rolling mean with window=50", "aapl['SMA_50'] = aapl['close'].rolling(window=50).mean()")
            
            # Exercise 4.2
            if "plt.plot(___, ___, label='Close" in line:
                line = line.replace("plt.plot(___, ___, label='Close Price', color='blue', alpha=0.6, linewidth=1)", "plt.plot(aapl['date'], aapl['close'], label='Close Price', color='blue', alpha=0.6, linewidth=1)")
            if "plt.plot(___, ___, label='SMA 20'" in line:
                line = line.replace("plt.plot(___, ___, label='SMA 20', color='orange', linewidth=2)", "plt.plot(aapl['date'], aapl['SMA_20'], label='SMA 20', color='orange', linewidth=2)")
            if "plt.plot(___, ___, label='SMA 50'" in line:
                line = line.replace("plt.plot(___, ___, label='SMA 50', color='green', linewidth=2)", "plt.plot(aapl['date'], aapl['SMA_50'], label='SMA 50', color='green', linewidth=2)")
            
            # Exercise 4.3
            if "aapl['signal'] = np.where(___, 1, 0)" in line:
                line = line.replace("aapl['signal'] = np.where(___, 1, 0)  # SMA_20 > SMA_50 condition", "aapl['signal'] = np.where(aapl['SMA_20'] > aapl['SMA_50'], 1, 0)")
            if "plt.scatter(___, ___, marker='^', color='green', s=150, label='Golden Cross'" in line:
                line = line.replace("plt.scatter(___, ___, marker='^', color='green', s=150, label='Golden Cross', zorder=5)", "plt.scatter(golden_cross['date'], golden_cross['close'], marker='^', color='green', s=150, label='Golden Cross', zorder=5)")
            if "plt.scatter(___, ___, marker='v', color='red', s=150, label='Death Cross'" in line:
                line = line.replace("plt.scatter(___, ___, marker='v', color='red', s=150, label='Death Cross', zorder=5)", "plt.scatter(death_cross['date'], death_cross['close'], marker='v', color='red', s=150, label='Death Cross', zorder=5)")
            
            # Exercise 4.4
            if "aapl['EMA_20'] = ___" in line:
                line = line.replace("aapl['EMA_20'] = ___  # ewm(span=20, adjust=False).mean()", "aapl['EMA_20'] = aapl['close'].ewm(span=20, adjust=False).mean()")
            
            # Exercise 5.1
            if "price_pivot = ___  # pivot_table" in line:
                line = line.replace("price_pivot = ___  # pivot_table with close prices", "price_pivot = df.pivot_table(values='close', index='date', columns='ticker')")
            
            # Exercise 5.2
            if "returns_pivot = ___  # pct_change" in line:
                line = line.replace("returns_pivot = ___  # pct_change() on price_pivot", "returns_pivot = price_pivot.pct_change()")
            if "correlation_matrix = ___  # .corr" in line:
                line = line.replace("correlation_matrix = ___  # .corr() on returns_pivot", "correlation_matrix = returns_pivot.corr()")
            
            # Exercise 5.3 Heatmap
            if "___,              # correlation matrix" in line:
                line = line.replace("___,              # correlation matrix", "    correlation_matrix,")
            
            # Exercise 5.4 Sharpe
            if "annual_return = ___      # mean_daily" in line:
                line = line.replace("annual_return = ___      # mean_daily * 252", "annual_return = mean_daily * 252")
            if "annual_volatility = ___  # std_daily" in line:
                line = line.replace("annual_volatility = ___  # std_daily * sqrt(252)", "annual_volatility = std_daily * np.sqrt(252)")
            if "sharpe_ratio = ___  # (annual_return" in line:
                line = line.replace("sharpe_ratio = ___  # (annual_return - risk_free_rate) / annual_volatility", "sharpe_ratio = (annual_return - risk_free_rate) / annual_volatility")
            
            # Exercise 6.1 Candlestick
            if "ax.bar(up.index, ___, width, bottom=___, color='green'" in line:
                line = line.replace("ax.bar(up.index, ___, width, bottom=___, color='green', edgecolor='black', linewidth=0.5)  # height = close-open, bottom = open", "ax.bar(up.index, up['close'] - up['open'], width, bottom=up['open'], color='green', edgecolor='black', linewidth=0.5)")
            if "ax.vlines(up.index, ___, ___, color='black'" in line:
                line = line.replace("ax.vlines(up.index, ___, ___, color='black', linewidth=0.8)  # low to high", "ax.vlines(up.index, up['low'], up['high'], color='black', linewidth=0.8)")
            if "ax.bar(down.index, ___, width, bottom=___, color='red'" in line:
                line = line.replace("ax.bar(down.index, ___, width, bottom=___, color='red', edgecolor='black', linewidth=0.5)  # height = open-close, bottom = close", "ax.bar(down.index, down['open'] - down['close'], width, bottom=down['close'], color='red', edgecolor='black', linewidth=0.5)")
            if "ax.vlines(down.index, ___, ___, color='black'" in line:
                line = line.replace("ax.vlines(down.index, ___, ___, color='black', linewidth=0.8)  # low to high", "ax.vlines(down.index, down['low'], down['high'], color='black', linewidth=0.8)")
            
            # Exercise 6.2 Rolling Volatility
            if "rolling_vol = ___  # rolling(window=30)" in line:
                line = line.replace("rolling_vol = ___  # rolling(window=30).std() on daily_return", "rolling_vol = stock_data['daily_return'].rolling(window=30).std()")
            if "annualized_rolling_vol = ___  # rolling_vol" in line:
                line = line.replace("annualized_rolling_vol = ___  # rolling_vol * sqrt(252)", "annualized_rolling_vol = rolling_vol * np.sqrt(252)")
            
            new_source.append(line)
        cell["source"] = new_source

with open("solutions/project_3_stock_trends_solution.ipynb", "w", encoding="utf-8") as f:
    json.dump(nb3, f, indent=2, ensure_ascii=False)


# ============================================================
# SOLVE EXCEL WORKBOOK
# ============================================================
# Create a completed Excel workbook with all answers filled
print("Solving Project 4 (Excel Mastery)...")

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = load_workbook("project_4_excel_mastery.xlsx")
    
    # 1. Fill Dynamic Arrays (Simulation of student's completed dynamic array sheets)
    ws_da = wb["Dynamic_Arrays"]
    # We will write some simulated formulas and labels to show how the solved version works
    ws_da.cell(row=4, column=2, value="=UNIQUE(Raw_Sales_Data[REGION])").font = Font(name="Consolas", size=10, italic=True)
    ws_da.cell(row=11, column=2, value="=FILTER(Raw_Sales_Data, (Raw_Sales_Data[CATEGORY]=\"Software\")*(Raw_Sales_Data[REVENUE]>5000))").font = Font(name="Consolas", size=10, italic=True)
    ws_da.cell(row=23, column=2, value="=SORT(UNIQUE(Raw_Sales_Data[PRODUCT]))").font = Font(name="Consolas", size=10, italic=True)
    ws_da.cell(row=35, column=2, value="=SORTBY(UNIQUE(Raw_Sales_Data[SALES_REP]), SUMIF(Raw_Sales_Data[SALES_REP], UNIQUE(Raw_Sales_Data[SALES_REP]), Raw_Sales_Data[PROFIT]))").font = Font(name="Consolas", size=10, italic=True)
    ws_da.cell(row=47, column=2, value="=LET(rev, SUMIFS(Raw_Sales_Data[REVENUE], Raw_Sales_Data[REGION], B4), prof, SUMIFS(Raw_Sales_Data[PROFIT], Raw_Sales_Data[REGION], B4), prof/rev)").font = Font(name="Consolas", size=10, italic=True)
    ws_da.cell(row=57, column=2, value="=XLOOKUP(MAX(SUMIFS(Raw_Sales_Data[REVENUE], Raw_Sales_Data[SALES_REP], UNIQUE(Raw_Sales_Data[SALES_REP]))), SUMIFS(Raw_Sales_Data[REVENUE], Raw_Sales_Data[SALES_REP], UNIQUE(Raw_Sales_Data[SALES_REP])), UNIQUE(Raw_Sales_Data[SALES_REP]))").font = Font(name="Consolas", size=10, italic=True)
    
    # 2. Fill Conditional Analysis SUMIFS/COUNTIFS formulas
    ws_ca = wb["Conditional_Analysis"]
    # Loop over the regions and write the actual SUMIFS formula
    for row in range(5, 10):
        region = ws_ca.cell(row=row, column=1).value
        # Write exact Excel formula
        formula_cell = ws_ca.cell(row=row, column=2, value=f"=SUMIFS(Raw_Sales_Data!J:J, Raw_Sales_Data!C:C, A{row})")
        formula_cell.font = Font(name="Calibri", size=11)
        formula_cell.number_format = '$#,##0.00'
        
    # Write COUNTIFS formulas
    categories = ["Software", "Hardware", "Services", "Accessories"]
    regions = ["North", "South", "East", "West", "Central"]
    for i, cat in enumerate(categories, 14):
        for j, region in enumerate(regions, 2):
            col_letter = get_column_letter(j)
            formula_cell = ws_ca.cell(row=i, column=j, value=f"=COUNTIFS(Raw_Sales_Data!$E:$E, $A{i}, Raw_Sales_Data!$C:$C, {col_letter}$13)")
            formula_cell.font = Font(name="Calibri", size=11)
            formula_cell.number_format = '#,##0'
            
    # Save the solved version
    wb.save("solutions/project_4_excel_mastery_solution.xlsx")
    print("Excel solved workbook generated successfully!")
except Exception as e:
    print(f"Error solving Excel: {e}")

# ============================================================
# EXECUTE NOTEBOOKS (using nbconvert)
# ============================================================
print("Executing notebooks to populate outputs...")
notebooks_to_run = [
    "solutions/project_1_ecommerce_solution.ipynb",
    "solutions/project_2_health_risk_solution.ipynb",
    "solutions/project_3_stock_trends_solution.ipynb"
]

for nb_path in notebooks_to_run:
    print(f"Executing {nb_path}...")
    try:
        # Run using python -m jupyter nbconvert
        subprocess.check_call([
            "python", "-m", "jupyter", "nbconvert", 
            "--to", "notebook", 
            "--execute", 
            "--ExecutePreprocessor.timeout=180", 
            "--inplace", 
            nb_path
        ])
        print(f"Successfully executed and saved {nb_path}!")
    except Exception as e:
        print(f"Warning: Could not run execution for {nb_path} (it is still saved with the correct code): {str(e)}")

print("Solutions generated successfully!")

